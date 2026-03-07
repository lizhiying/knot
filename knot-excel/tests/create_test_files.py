import openpyxl
import os

base = os.path.dirname(os.path.abspath(__file__))

# 场景1：4行空行分隔
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'MultiBlock'
ws['A1'] = '姓名'; ws['B1'] = '部门'; ws['C1'] = '职级'
for i, (a, b, c) in enumerate([('张三','技术部','P7'), ('李四','市场部','P6'), ('王五','技术部','P8')]):
    ws.cell(row=2+i, column=1, value=a); ws.cell(row=2+i, column=2, value=b); ws.cell(row=2+i, column=3, value=c)
ws['A9'] = '月份'; ws['B9'] = '销售额'; ws['C9'] = '利润'
for i, (a, b, c) in enumerate([('1月',10000,3000), ('2月',12000,3500), ('3月',15000,4200), ('4月',11000,3100)]):
    ws.cell(row=10+i, column=1, value=a); ws.cell(row=10+i, column=2, value=b); ws.cell(row=10+i, column=3, value=c)
wb.save(os.path.join(base, 'test_multi_block.xlsx'))

# 场景2：类型跳变（只有1行空行 + 类型跳变）
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = 'TypeTransition'
ws2['A1'] = '产品'; ws2['B1'] = '价格'; ws2['C1'] = '库存'
for i, (a, b, c) in enumerate([('产品A',100,500), ('产品B',200,300), ('产品C',150,400)]):
    ws2.cell(row=2+i, column=1, value=a); ws2.cell(row=2+i, column=2, value=b); ws2.cell(row=2+i, column=3, value=c)
ws2['A6'] = '供应商'; ws2['B6'] = '联系人'; ws2['C6'] = '评分'
for i, (a, b, c) in enumerate([('供应商X','张经理',95), ('供应商Y','李总',88), ('供应商Z','王主任',91)]):
    ws2.cell(row=7+i, column=1, value=a); ws2.cell(row=7+i, column=2, value=b); ws2.cell(row=7+i, column=3, value=c)
wb2.save(os.path.join(base, 'test_type_transition.xlsx'))

# 场景3：数据中间有空行（不应该被分割）
wb3 = openpyxl.Workbook()
ws3 = wb3.active
ws3.title = 'WithGaps'
ws3['A1'] = '项目'; ws3['B1'] = '金额'; ws3['C1'] = '状态'
ws3.cell(row=2,column=1,value='项目A'); ws3.cell(row=2,column=2,value=50000); ws3.cell(row=2,column=3,value='进行中')
ws3.cell(row=3,column=1,value='项目B'); ws3.cell(row=3,column=2,value=80000); ws3.cell(row=3,column=3,value='已完成')
ws3.cell(row=5,column=1,value='项目C'); ws3.cell(row=5,column=2,value=120000); ws3.cell(row=5,column=3,value='进行中')
ws3.cell(row=6,column=1,value='项目D'); ws3.cell(row=6,column=2,value=30000); ws3.cell(row=6,column=3,value='已暂停')
ws3.cell(row=8,column=1,value='项目E'); ws3.cell(row=8,column=2,value=90000); ws3.cell(row=8,column=3,value='已完成')
wb3.save(os.path.join(base, 'test_with_gaps.xlsx'))

print('Done!')
